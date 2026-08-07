"""Structured Phase 03 model evidence and exact scientific handoff resolution."""

from __future__ import annotations

import json
from collections.abc import Sequence
from datetime import UTC, datetime
from enum import StrEnum
from pathlib import Path
from typing import Annotated, Literal, NamedTuple, Protocol, Self

from pydantic import (
    BaseModel,
    ConfigDict,
    Field,
    StringConstraints,
    model_validator,
)

from buduunkhad.ai.fingerprint import sha256_value
from buduunkhad.core.boundary_validation import load_boundary_acceptance
from buduunkhad.core.evidence_manifest import (
    EvidenceAuthorityResolver,
    EvidenceLifecycleState,
    EvidenceManifestBinding,
    EvidenceRole,
    ResolvedEvidence,
    evidence_bindings,
)
from buduunkhad.core.execution_policy import ExecutionMode
from buduunkhad.core.georeference import load_georeference_acceptance
from buduunkhad.core.run_artifacts import sha256_file
from buduunkhad.core.run_storage import SourcePhaseBinding, resolve_source_phase, validate_run_id

DEPOSIT_MODEL_ASSESSMENT_FORMAT_VERSION = "1.1.0"
DEPOSIT_MODEL_CRITIQUE_FORMAT_VERSION = "1.0.0"
DEPOSIT_MODEL_REVIEW_FORMAT_VERSION = "1.0.0"
PHASE03_HANDOFF_FORMAT_VERSION = "1.0.0"
PHASE03_HANDOFF_RESOLVER = "buduunkhad.phase03.scientific-handoff-resolver-v1"
PHASE03_REQUIRED_EVIDENCE_ROLES = frozenset(
    {
        EvidenceRole.GEOLOGY,
        EvidenceRole.STRUCTURE,
        EvidenceRole.OCCURRENCE,
        EvidenceRole.ALTERATION_SUPPORT,
    }
)

DEPOSIT_MODEL_SCORING_CRITERIA: tuple[tuple[str, str, int], ...] = (
    ("favorable-geology", "Favorable geology / host lithology", 20),
    ("structure-contact", "Intrusive / contact / structure control", 15),
    ("known-occurrence", "Known mineral occurrence", 15),
    (
        "historical-geochemistry",
        "Historical geochemistry / shlich / stream sediment",
        15,
    ),
    ("metallogenic-context", "Metallogenic context", 10),
    ("remote-sensing-alteration", "ASTER/Sentinel alteration support", 10),
    ("field-pxrf", "Field mapping / pXRF support", 10),
    ("access-workability", "Access / workability", 5),
)

# How diagnostic each rubric criterion is for each candidate model, as a share of that criterion's
# maximum. A structurally controlled vein earns the full structure weight; a placer earns little
# from it. These shares express which evidence class discriminates which model — they are not
# probabilities, and they never create evidence: a criterion with nothing assembled scores zero
# for every model regardless of its share.
DEPOSIT_MODEL_CRITERION_AFFINITY: dict[str, dict[str, float]] = {
    "Au-Cu hydrothermal vein": {
        "favorable-geology": 0.8,
        "structure-contact": 1.0,
        "known-occurrence": 1.0,
        "historical-geochemistry": 1.0,
        "metallogenic-context": 0.8,
        "remote-sensing-alteration": 0.8,
        "field-pxrf": 1.0,
        "access-workability": 1.0,
    },
    "Intrusion-related Cu-Au-Mo": {
        "favorable-geology": 1.0,
        "structure-contact": 0.9,
        "known-occurrence": 0.9,
        "historical-geochemistry": 1.0,
        "metallogenic-context": 0.9,
        "remote-sensing-alteration": 1.0,
        "field-pxrf": 1.0,
        "access-workability": 1.0,
    },
    "Skarn / contact metasomatic": {
        "favorable-geology": 1.0,
        "structure-contact": 1.0,
        "known-occurrence": 0.8,
        "historical-geochemistry": 0.8,
        "metallogenic-context": 0.7,
        "remote-sensing-alteration": 0.7,
        "field-pxrf": 1.0,
        "access-workability": 1.0,
    },
    "Polymetallic vein": {
        "favorable-geology": 0.7,
        "structure-contact": 1.0,
        "known-occurrence": 0.9,
        "historical-geochemistry": 1.0,
        "metallogenic-context": 0.7,
        "remote-sensing-alteration": 0.5,
        "field-pxrf": 1.0,
        "access-workability": 1.0,
    },
    "VMS possibility": {
        "favorable-geology": 0.9,
        "structure-contact": 0.5,
        "known-occurrence": 0.6,
        "historical-geochemistry": 0.8,
        "metallogenic-context": 0.6,
        "remote-sensing-alteration": 0.5,
        "field-pxrf": 1.0,
        "access-workability": 1.0,
    },
    "Heavy mineral / placer": {
        "favorable-geology": 0.4,
        "structure-contact": 0.2,
        "known-occurrence": 0.5,
        "historical-geochemistry": 1.0,
        "metallogenic-context": 0.4,
        "remote-sensing-alteration": 0.3,
        "field-pxrf": 0.8,
        "access-workability": 1.0,
    },
}

# Pathfinder elements that discriminate between models. Applied only to the geochemistry criterion
# and only to elements actually measured in the assembled anomaly attributes.
DEPOSIT_MODEL_ELEMENT_AFFINITY: dict[str, frozenset[str]] = {
    "Au-Cu hydrothermal vein": frozenset({"au", "ag", "cu", "as", "sb", "te"}),
    "Intrusion-related Cu-Au-Mo": frozenset({"cu", "mo", "au", "w", "bi", "re"}),
    "Skarn / contact metasomatic": frozenset({"w", "sn", "cu", "zn", "fe", "bi", "mo"}),
    "Polymetallic vein": frozenset({"pb", "zn", "ag", "ba", "cu", "sb", "mn"}),
    "VMS possibility": frozenset({"cu", "zn", "pb", "ni", "co", "ba", "pyrite"}),
    "Heavy mineral / placer": frozenset({"au", "sn", "ti", "zr", "cr", "pt", "w"}),
}


class CriterionEvidence(NamedTuple):
    """What one 03A criterion actually has behind it after a run."""

    criterion_id: str
    criterion: str
    max_points: int
    evidence_ids: tuple[str, ...]
    feature_count: int
    detail: str

    @property
    def present(self) -> bool:
        return bool(self.evidence_ids) and self.feature_count > 0


def score_deposit_models(
    evidence: Sequence[CriterionEvidence],
    *,
    observed_elements: Sequence[str] = (),
) -> dict[str, tuple[DepositModelCriterionScore, ...]]:
    """Score every candidate model against the rubric using only assembled evidence.

    Deterministic and reproducible: the same evidence always yields the same matrix. A criterion
    with no evidence is reported as missing at zero rather than estimated, so an unscored model is
    visibly unscored instead of quietly averaged.
    """

    measured = {value.strip().casefold() for value in observed_elements if value.strip()}
    scored: dict[str, tuple[DepositModelCriterionScore, ...]] = {}
    for model, affinity in DEPOSIT_MODEL_CRITERION_AFFINITY.items():
        rows: list[DepositModelCriterionScore] = []
        for item in evidence:
            share = affinity.get(item.criterion_id, 0.0)
            if not item.present or share <= 0:
                rows.append(
                    DepositModelCriterionScore(
                        criterion_id=item.criterion_id,
                        criterion=item.criterion,
                        max_points=item.max_points,
                        proposed_points=0,
                        rationale=(
                            f"No evidence assembled for this criterion ({item.detail})."
                            if not item.present
                            else f"{item.criterion} is not diagnostic for {model}."
                        ),
                        missing=not item.present,
                    )
                )
                continue
            weight = share
            note = ""
            if item.criterion_id == "historical-geochemistry" and measured:
                pathfinders = DEPOSIT_MODEL_ELEMENT_AFFINITY.get(model, frozenset())
                matched = sorted(measured & pathfinders)
                weight = share * (len(matched) / len(measured))
                note = (
                    f" Pathfinders matching {model}: {', '.join(matched)} of "
                    f"{len(measured)} measured."
                    if matched
                    else f" No measured element is a {model} pathfinder."
                )
            points = round(item.max_points * weight, 1)
            rows.append(
                DepositModelCriterionScore(
                    criterion_id=item.criterion_id,
                    criterion=item.criterion,
                    max_points=item.max_points,
                    proposed_points=points,
                    evidence_ids=tuple(sorted(set(item.evidence_ids))) if points > 0 else (),
                    rationale=(
                        f"{item.detail}; {item.criterion} carries {share:.0%} of its weight for "
                        f"{model}.{note}"
                    ),
                    missing=False,
                )
            )
        scored[model] = tuple(rows)
    return scored


def deposit_model_total(scores: Sequence[DepositModelCriterionScore]) -> float:
    return round(sum(item.proposed_points for item in scores), 1)


Sha256 = Annotated[str, StringConstraints(pattern=r"^[0-9a-f]{64}$")]
NonEmpty = Annotated[str, StringConstraints(strip_whitespace=True, min_length=1)]


class Phase03ScienceError(RuntimeError):
    """A Phase 03 scientific record or one of its authorities cannot be trusted."""


class EvidenceResolver(Protocol):
    """Minimum authority-resolver interface needed to create a 03A assessment."""

    def resolve_selected(
        self,
        manifest_ids: tuple[str, ...] | list[str],
    ) -> tuple[ResolvedEvidence, ...]: ...


class _StrictModel(BaseModel):
    model_config = ConfigDict(
        extra="forbid",
        frozen=True,
        revalidate_instances="always",
        allow_inf_nan=False,
    )

    @classmethod
    def model_construct(
        cls,
        _fields_set: set[str] | None = None,
        **values: object,
    ) -> Self:
        del _fields_set, values
        raise TypeError("model_construct is unsupported; use validated construction")


class EvidenceRelationship(StrEnum):
    SUPPORTING = "supporting"
    CONTRADICTING = "contradicting"


class DepositModelEvidenceGroup(_StrictModel):
    """Correlated evidence is grouped so one derivative family is counted once."""

    group_id: NonEmpty
    relationship: EvidenceRelationship
    evidence_ids: tuple[NonEmpty, ...] = Field(min_length=1)
    source_scale: NonEmpty
    source_date: NonEmpty
    correlation_basis: NonEmpty
    counted_once: Literal[True] = True
    confidence_basis: NonEmpty
    limitations: tuple[NonEmpty, ...] = ()

    @model_validator(mode="after")
    def _ordered_group(self) -> DepositModelEvidenceGroup:
        if tuple(sorted(set(self.evidence_ids))) != self.evidence_ids:
            raise ValueError("deposit-model evidence IDs must be unique and ordered")
        if tuple(sorted(set(self.limitations))) != self.limitations:
            raise ValueError("deposit-model group limitations must be unique and ordered")
        return self


class DepositModelCriterionScore(_StrictModel):
    """One explicit row of the adopted 100-point 03A rubric."""

    criterion_id: NonEmpty
    criterion: NonEmpty
    max_points: int = Field(gt=0)
    proposed_points: float = Field(ge=0)
    evidence_ids: tuple[NonEmpty, ...] = ()
    rationale: NonEmpty
    missing: bool

    @model_validator(mode="after")
    def _truthful_score(self) -> DepositModelCriterionScore:
        if self.proposed_points > self.max_points:
            raise ValueError("deposit-model criterion score exceeds its maximum")
        if tuple(sorted(set(self.evidence_ids))) != self.evidence_ids:
            raise ValueError("criterion evidence IDs must be unique and ordered")
        if self.missing and (self.proposed_points != 0 or self.evidence_ids):
            raise ValueError("missing deposit-model criteria must score zero with no evidence")
        if self.proposed_points > 0 and not self.evidence_ids:
            raise ValueError("positive deposit-model criterion scores require exact evidence IDs")
        return self


class _DepositModelAssessmentIdentity(_StrictModel):
    format_version: Literal["1.0.0", "1.1.0"] = DEPOSIT_MODEL_ASSESSMENT_FORMAT_VERSION
    phase03_run_id: NonEmpty
    candidate_model: NonEmpty
    evidence_manifest_bindings: tuple[EvidenceManifestBinding, ...] = Field(min_length=1)
    evidence_groups: tuple[DepositModelEvidenceGroup, ...] = Field(min_length=1)
    selected_evidence_ids: tuple[NonEmpty, ...] = Field(min_length=1)
    missing_evidence: tuple[NonEmpty, ...]
    confidence_basis: NonEmpty
    limitations: tuple[NonEmpty, ...] = Field(min_length=1)
    recommended_validation: tuple[NonEmpty, ...] = Field(min_length=1)
    criterion_scores: tuple[DepositModelCriterionScore, ...] = ()
    draft_score: float | None = Field(default=None, ge=0, le=100)
    proposing_job_id: NonEmpty | None = None
    proposing_response_id: NonEmpty | None = None
    component: Literal["buduunkhad.phase03.deposit-model-assessment-v1"] = (
        "buduunkhad.phase03.deposit-model-assessment-v1"
    )

    @model_validator(mode="after")
    def _coherent_assessment(self) -> _DepositModelAssessmentIdentity:
        validate_run_id(self.phase03_run_id)
        manifest_ids = tuple(item.manifest_id for item in self.evidence_manifest_bindings)
        if tuple(sorted(set(manifest_ids))) != manifest_ids:
            raise ValueError("deposit-model evidence bindings must be unique and ordered")
        group_ids = tuple(item.group_id for item in self.evidence_groups)
        if tuple(sorted(set(group_ids))) != group_ids:
            raise ValueError("deposit-model evidence groups must be unique and ordered")
        grouped = tuple(
            sorted(
                evidence_id for group in self.evidence_groups for evidence_id in group.evidence_ids
            )
        )
        if (
            tuple(sorted(set(self.selected_evidence_ids))) != self.selected_evidence_ids
            or grouped != self.selected_evidence_ids
        ):
            raise ValueError(
                "deposit-model selected evidence must exactly match unique grouped evidence"
            )
        if not any(
            item.relationship is EvidenceRelationship.SUPPORTING for item in self.evidence_groups
        ):
            raise ValueError("deposit-model assessment requires supporting evidence")
        if (self.proposing_job_id is None) != (self.proposing_response_id is None):
            raise ValueError("AI proposal job and response identities must be paired")
        if self.format_version == "1.0.0":
            if self.criterion_scores:
                raise ValueError("legacy deposit-model assessments cannot contain criterion scores")
        else:
            expected = tuple(item[0] for item in DEPOSIT_MODEL_SCORING_CRITERIA)
            actual = tuple(item.criterion_id for item in self.criterion_scores)
            if actual != expected:
                raise ValueError(
                    "deposit-model assessment requires the exact eight-criterion rubric"
                )
            for score, (criterion_id, criterion, maximum) in zip(
                self.criterion_scores,
                DEPOSIT_MODEL_SCORING_CRITERIA,
                strict=True,
            ):
                if (
                    score.criterion_id != criterion_id
                    or score.criterion != criterion
                    or score.max_points != maximum
                ):
                    raise ValueError("deposit-model criterion identity or maximum is incorrect")
                if not set(score.evidence_ids) <= set(self.selected_evidence_ids):
                    raise ValueError("criterion score cites evidence outside the assessment")
            total = sum(item.proposed_points for item in self.criterion_scores)
            if self.draft_score is None or not abs(total - self.draft_score) <= 1e-9:
                raise ValueError("deposit-model draft score must equal the criterion total")
        for values, description in (
            (self.missing_evidence, "missing evidence"),
            (self.limitations, "limitations"),
            (self.recommended_validation, "recommended validation"),
        ):
            if tuple(sorted(set(values))) != values:
                raise ValueError(f"deposit-model {description} must be unique and ordered")
        return self


class DepositModelAssessment(_DepositModelAssessmentIdentity):
    """A structured proposal; it contains no geologist acceptance."""

    assessment_id: Sha256

    @model_validator(mode="after")
    def _sealed_identity(self) -> DepositModelAssessment:
        identity = _DepositModelAssessmentIdentity.model_validate(
            self.model_dump(mode="python", exclude={"assessment_id"})
        )
        if self.assessment_id != sha256_value(identity):
            raise ValueError("deposit-model assessment identity is invalid")
        return self

    @classmethod
    def create(cls, **values: object) -> DepositModelAssessment:
        identity = _DepositModelAssessmentIdentity.model_validate(values)
        return cls(
            **identity.model_dump(mode="python"),
            assessment_id=sha256_value(identity),
        )


class CritiqueOrigin(StrEnum):
    AI_INDEPENDENT = "ai-independent"
    HUMAN_INDEPENDENT = "human-independent"


class CritiqueConclusion(StrEnum):
    SUPPORTED = "supported"
    CHALLENGED = "challenged"


class _DepositModelCritiqueIdentity(_StrictModel):
    format_version: Literal["1.0.0"] = DEPOSIT_MODEL_CRITIQUE_FORMAT_VERSION
    assessment_id: Sha256
    assessment_file_sha256: Sha256
    phase03_run_id: NonEmpty
    critic: NonEmpty
    origin: CritiqueOrigin
    critic_authorization_id: NonEmpty | None = None
    critiqued_at: datetime
    critique_job_id: NonEmpty | None = None
    critique_response_id: NonEmpty | None = None
    conclusion: CritiqueConclusion
    findings: tuple[NonEmpty, ...] = Field(min_length=1)
    limitations: tuple[NonEmpty, ...] = ()

    @model_validator(mode="after")
    def _independent_critique(self) -> _DepositModelCritiqueIdentity:
        validate_run_id(self.phase03_run_id)
        if self.critiqued_at.tzinfo is None or self.critiqued_at.utcoffset() is None:
            raise ValueError("critique time must be timezone-aware")
        if self.critiqued_at.utcoffset() != UTC.utcoffset(None):
            raise ValueError("critique time must be recorded in UTC")
        ai = self.origin is CritiqueOrigin.AI_INDEPENDENT
        if ai != bool(self.critique_job_id and self.critique_response_id):
            raise ValueError("AI critique requires its separate job and response identities")
        if ai and self.critic_authorization_id is not None:
            raise ValueError("AI critique cannot claim a human authorization identity")
        if not ai and (
            self.critic_authorization_id is None
            or self.critique_job_id is not None
            or self.critique_response_id is not None
        ):
            raise ValueError("human critique requires authorization and no AI job identity")
        if tuple(sorted(set(self.findings))) != self.findings:
            raise ValueError("critique findings must be unique and ordered")
        if tuple(sorted(set(self.limitations))) != self.limitations:
            raise ValueError("critique limitations must be unique and ordered")
        return self


class DepositModelCritique(_DepositModelCritiqueIdentity):
    critique_id: Sha256

    @model_validator(mode="after")
    def _sealed_identity(self) -> DepositModelCritique:
        identity = _DepositModelCritiqueIdentity.model_validate(
            self.model_dump(mode="python", exclude={"critique_id"})
        )
        if self.critique_id != sha256_value(identity):
            raise ValueError("deposit-model critique identity is invalid")
        return self

    @classmethod
    def create(cls, **values: object) -> DepositModelCritique:
        identity = _DepositModelCritiqueIdentity.model_validate(values)
        return cls(
            **identity.model_dump(mode="python"),
            critique_id=sha256_value(identity),
        )


class ScientificReviewDecision(StrEnum):
    ACCEPTED = "accepted"
    REJECTED = "rejected"


class _DepositModelReviewIdentity(_StrictModel):
    format_version: Literal["1.0.0"] = DEPOSIT_MODEL_REVIEW_FORMAT_VERSION
    assessment_id: Sha256
    assessment_file_sha256: Sha256
    critique_id: Sha256
    critique_file_sha256: Sha256
    phase03_run_id: NonEmpty
    reviewer: NonEmpty
    reviewer_role: Literal["project-geologist"] = "project-geologist"
    reviewer_authorization_id: NonEmpty
    reviewed_at: datetime
    decision: ScientificReviewDecision
    accepted_model: NonEmpty | None
    accepted_confidence: NonEmpty | None
    rationale: NonEmpty
    limitations: tuple[NonEmpty, ...] = ()

    @model_validator(mode="after")
    def _truthful_review(self) -> _DepositModelReviewIdentity:
        validate_run_id(self.phase03_run_id)
        if self.reviewed_at.tzinfo is None or self.reviewed_at.utcoffset() is None:
            raise ValueError("deposit-model review time must be timezone-aware")
        if self.reviewed_at.utcoffset() != UTC.utcoffset(None):
            raise ValueError("deposit-model review time must be recorded in UTC")
        accepted = self.decision is ScientificReviewDecision.ACCEPTED
        if accepted != bool(self.accepted_model and self.accepted_confidence):
            raise ValueError("accepted model and confidence exist only for an accepted review")
        if tuple(sorted(set(self.limitations))) != self.limitations:
            raise ValueError("deposit-model review limitations must be unique and ordered")
        return self


class DepositModelReview(_DepositModelReviewIdentity):
    review_id: Sha256

    @model_validator(mode="after")
    def _sealed_identity(self) -> DepositModelReview:
        identity = _DepositModelReviewIdentity.model_validate(
            self.model_dump(mode="python", exclude={"review_id"})
        )
        if self.review_id != sha256_value(identity):
            raise ValueError("deposit-model review identity is invalid")
        return self

    @classmethod
    def create(cls, **values: object) -> DepositModelReview:
        identity = _DepositModelReviewIdentity.model_validate(values)
        return cls(**identity.model_dump(mode="python"), review_id=sha256_value(identity))


class GeoreferenceAcceptanceBinding(_StrictModel):
    acceptance_id: Sha256
    acceptance_file_sha256: Sha256
    source_sha256: Sha256
    derivative_sha256: Sha256


class ScientificRecordBinding(_StrictModel):
    record_id: Sha256
    file_sha256: Sha256


class _Phase03ScientificHandoffIdentity(_StrictModel):
    format_version: Literal["1.0.0"] = PHASE03_HANDOFF_FORMAT_VERSION
    phase03_source: SourcePhaseBinding
    boundary_acceptance: ScientificRecordBinding
    georeference_acceptances: tuple[GeoreferenceAcceptanceBinding, ...] = Field(min_length=1)
    evidence_manifest_bindings: tuple[EvidenceManifestBinding, ...] = Field(min_length=1)
    accepted_evidence_ids: tuple[NonEmpty, ...] = Field(min_length=1)
    accepted_evidence_roles: tuple[EvidenceRole, ...] = Field(min_length=1)
    deposit_model_assessment: ScientificRecordBinding
    deposit_model_critique: ScientificRecordBinding
    deposit_model_review: ScientificRecordBinding
    unresolved_gaps: tuple[NonEmpty, ...]
    limitations: tuple[NonEmpty, ...] = Field(min_length=1)
    handed_off_by: NonEmpty
    handoff_authorization_id: NonEmpty
    handed_off_at: datetime
    decision: Literal["accepted-for-phase04-input"]
    readiness_ids_resolved: tuple[
        Literal["METH-READY-004"],
        Literal["METH-READY-005"],
        Literal["METH-READY-006"],
    ] = ("METH-READY-004", "METH-READY-005", "METH-READY-006")
    resolver_component: Literal["buduunkhad.phase03.scientific-handoff-resolver-v1"] = (
        PHASE03_HANDOFF_RESOLVER
    )

    @model_validator(mode="after")
    def _coherent_handoff(self) -> _Phase03ScientificHandoffIdentity:
        if self.phase03_source.phase_id != "03":
            raise ValueError("scientific handoff must bind a sealed Phase 03 source")
        georef_ids = tuple(item.acceptance_id for item in self.georeference_acceptances)
        if tuple(sorted(set(georef_ids))) != georef_ids:
            raise ValueError("georeference acceptances must be unique and ordered")
        manifest_ids = tuple(item.manifest_id for item in self.evidence_manifest_bindings)
        if tuple(sorted(set(manifest_ids))) != manifest_ids:
            raise ValueError("evidence manifest bindings must be unique and ordered")
        if tuple(sorted(set(self.accepted_evidence_ids))) != self.accepted_evidence_ids:
            raise ValueError("accepted evidence IDs must be unique and ordered")
        if tuple(sorted(set(self.accepted_evidence_roles), key=lambda item: item.value)) != (
            self.accepted_evidence_roles
        ):
            raise ValueError("accepted evidence roles must be unique and ordered")
        if not set(self.accepted_evidence_roles) >= PHASE03_REQUIRED_EVIDENCE_ROLES:
            raise ValueError("scientific handoff lacks one or more required evidence roles")
        if tuple(sorted(set(self.unresolved_gaps))) != self.unresolved_gaps:
            raise ValueError("handoff gaps must be unique and ordered")
        if tuple(sorted(set(self.limitations))) != self.limitations:
            raise ValueError("handoff limitations must be unique and ordered")
        if self.handed_off_at.tzinfo is None or self.handed_off_at.utcoffset() is None:
            raise ValueError("handoff time must be timezone-aware")
        if self.handed_off_at.utcoffset() != UTC.utcoffset(None):
            raise ValueError("handoff time must be recorded in UTC")
        return self


class Phase03ScientificHandoff(_Phase03ScientificHandoffIdentity):
    handoff_id: Sha256

    @model_validator(mode="after")
    def _sealed_identity(self) -> Phase03ScientificHandoff:
        identity = _Phase03ScientificHandoffIdentity.model_validate(
            self.model_dump(mode="python", exclude={"handoff_id"})
        )
        if self.handoff_id != sha256_value(identity):
            raise ValueError("Phase 03 scientific handoff identity is invalid")
        return self

    @classmethod
    def create(cls, **values: object) -> Phase03ScientificHandoff:
        identity = _Phase03ScientificHandoffIdentity.model_validate(values)
        return cls(**identity.model_dump(mode="python"), handoff_id=sha256_value(identity))


def create_deposit_model_assessment(
    *,
    phase03_run_id: str,
    resolver: EvidenceResolver,
    evidence_manifest_ids: tuple[str, ...],
    candidate_model: str,
    evidence_groups: tuple[DepositModelEvidenceGroup, ...],
    missing_evidence: tuple[str, ...],
    confidence_basis: str,
    limitations: tuple[str, ...],
    recommended_validation: tuple[str, ...],
    criterion_scores: tuple[DepositModelCriterionScore, ...],
    draft_score: float | None = None,
    proposing_job_id: str | None = None,
    proposing_response_id: str | None = None,
) -> DepositModelAssessment:
    resolved = resolver.resolve_selected(evidence_manifest_ids)
    selected_ids = tuple(
        sorted(evidence_id for group in evidence_groups for evidence_id in group.evidence_ids)
    )
    by_id = {item.record.evidence_id: item for item in resolved}
    if len(by_id) != len(resolved) or not set(selected_ids) <= set(by_id):
        raise Phase03ScienceError("deposit-model evidence does not resolve exactly")
    for evidence_id in selected_ids:
        item = by_id[evidence_id]
        item.verify_current_artifact()
        if (
            item.record.lifecycle_state is not EvidenceLifecycleState.ACCEPTED_EVIDENCE
            or "03" not in item.record.eligible_phases
        ):
            raise Phase03ScienceError(
                f"deposit-model evidence is not accepted for Phase 03: {evidence_id}"
            )
    effective_score = (
        sum(item.proposed_points for item in criterion_scores)
        if draft_score is None
        else draft_score
    )
    return DepositModelAssessment.create(
        phase03_run_id=phase03_run_id,
        candidate_model=candidate_model,
        evidence_manifest_bindings=evidence_bindings(resolved),
        evidence_groups=tuple(sorted(evidence_groups, key=lambda item: item.group_id)),
        selected_evidence_ids=selected_ids,
        missing_evidence=tuple(sorted(set(missing_evidence))),
        confidence_basis=confidence_basis,
        limitations=tuple(sorted(set(limitations))),
        recommended_validation=tuple(sorted(set(recommended_validation))),
        criterion_scores=criterion_scores,
        draft_score=effective_score,
        proposing_job_id=proposing_job_id,
        proposing_response_id=proposing_response_id,
    )


def create_deposit_model_critique(
    assessment_path: Path,
    *,
    critic: str,
    origin: CritiqueOrigin,
    critiqued_at: datetime,
    conclusion: CritiqueConclusion,
    findings: tuple[str, ...],
    critic_authorization_id: str | None = None,
    critique_job_id: str | None = None,
    critique_response_id: str | None = None,
    limitations: tuple[str, ...] = (),
) -> DepositModelCritique:
    assessment = load_deposit_model_assessment(assessment_path)
    if assessment.proposing_job_id is not None and critique_job_id == assessment.proposing_job_id:
        raise Phase03ScienceError("independent critique must use a separate job identity")
    if (
        assessment.proposing_response_id is not None
        and critique_response_id == assessment.proposing_response_id
    ):
        raise Phase03ScienceError("independent critique must use a separate response identity")
    return DepositModelCritique.create(
        assessment_id=assessment.assessment_id,
        assessment_file_sha256=sha256_file(Path(assessment_path)),
        phase03_run_id=assessment.phase03_run_id,
        critic=critic,
        origin=origin,
        critic_authorization_id=critic_authorization_id,
        critiqued_at=critiqued_at,
        critique_job_id=critique_job_id,
        critique_response_id=critique_response_id,
        conclusion=conclusion,
        findings=tuple(sorted(set(findings))),
        limitations=tuple(sorted(set(limitations))),
    )


def create_deposit_model_review(
    assessment_path: Path,
    critique_path: Path,
    *,
    reviewer: str,
    reviewer_authorization_id: str,
    reviewed_at: datetime,
    decision: ScientificReviewDecision,
    accepted_model: str | None,
    accepted_confidence: str | None,
    rationale: str,
    limitations: tuple[str, ...] = (),
) -> DepositModelReview:
    assessment = load_deposit_model_assessment(assessment_path)
    critique = load_deposit_model_critique(critique_path)
    if (
        critique.assessment_id != assessment.assessment_id
        or critique.assessment_file_sha256 != sha256_file(Path(assessment_path))
        or critique.phase03_run_id != assessment.phase03_run_id
    ):
        raise Phase03ScienceError("deposit-model critique does not bind the exact assessment")
    return DepositModelReview.create(
        assessment_id=assessment.assessment_id,
        assessment_file_sha256=sha256_file(Path(assessment_path)),
        critique_id=critique.critique_id,
        critique_file_sha256=sha256_file(Path(critique_path)),
        phase03_run_id=assessment.phase03_run_id,
        reviewer=reviewer,
        reviewer_authorization_id=reviewer_authorization_id,
        reviewed_at=reviewed_at,
        decision=decision,
        accepted_model=accepted_model,
        accepted_confidence=accepted_confidence,
        rationale=rationale,
        limitations=tuple(sorted(set(limitations))),
    )


def resolve_phase03_scientific_handoff(
    *,
    runs_root: Path,
    evidence_root: Path,
    target_epsg: int,
    phase03_run_id: str,
    boundary_acceptance_path: Path,
    georeference_acceptance_paths: tuple[Path, ...],
    evidence_manifest_ids: tuple[str, ...],
    assessment_path: Path,
    critique_path: Path,
    review_path: Path,
    unresolved_gaps: tuple[str, ...],
    limitations: tuple[str, ...],
) -> Phase03ScientificHandoff:
    """Resolve real reviews over one sealed support-evidence Phase 03 run."""

    source = resolve_source_phase(
        runs_root,
        "03",
        phase03_run_id,
        require_advance=False,
        require_qaqc_passed=True,
    )
    if source.execution_mode is not ExecutionMode.SUPPORT_EVIDENCE:
        raise Phase03ScienceError("scientific handoff requires a support-evidence Phase 03 run")

    boundary = load_boundary_acceptance(boundary_acceptance_path)
    georeferences = tuple(
        load_georeference_acceptance(path) for path in georeference_acceptance_paths
    )
    if not georeferences or any(item.processing_run_id != phase03_run_id for item in georeferences):
        raise Phase03ScienceError(
            "georeference acceptances must bind the exact Phase 03 processing run"
        )

    resolver = EvidenceAuthorityResolver(
        runs_root=runs_root,
        evidence_root=evidence_root,
        target_epsg=target_epsg,
    )
    resolved = resolver.resolve_selected(evidence_manifest_ids)
    if not resolved:
        raise Phase03ScienceError("scientific handoff requires accepted evidence")
    for item in resolved:
        item.verify_current_artifact()
        if (
            item.record.lifecycle_state is not EvidenceLifecycleState.ACCEPTED_EVIDENCE
            or "03" not in item.record.eligible_phases
        ):
            raise Phase03ScienceError(
                f"scientific handoff evidence is not accepted for Phase 03: "
                f"{item.record.evidence_id}"
            )
    roles = tuple(
        sorted({item.record.evidence_role for item in resolved}, key=lambda item: item.value)
    )
    if not set(roles) >= PHASE03_REQUIRED_EVIDENCE_ROLES:
        missing = sorted(item.value for item in PHASE03_REQUIRED_EVIDENCE_ROLES - set(roles))
        raise Phase03ScienceError(f"scientific handoff lacks required evidence roles: {missing}")

    assessment = load_deposit_model_assessment(assessment_path)
    critique = load_deposit_model_critique(critique_path)
    review = load_deposit_model_review(review_path)
    assessment_hash = sha256_file(Path(assessment_path))
    critique_hash = sha256_file(Path(critique_path))
    selected_bindings = evidence_bindings(resolved)
    if (
        assessment.phase03_run_id != phase03_run_id
        or assessment.evidence_manifest_bindings != selected_bindings
        or not set(assessment.selected_evidence_ids)
        <= {item.record.evidence_id for item in resolved}
    ):
        raise Phase03ScienceError("deposit-model assessment does not bind the selected evidence")
    if (
        critique.assessment_id != assessment.assessment_id
        or critique.assessment_file_sha256 != assessment_hash
        or review.assessment_id != assessment.assessment_id
        or review.assessment_file_sha256 != assessment_hash
        or review.critique_id != critique.critique_id
        or review.critique_file_sha256 != critique_hash
        or review.phase03_run_id != phase03_run_id
    ):
        raise Phase03ScienceError("deposit-model review chain is inconsistent")
    if review.decision is not ScientificReviewDecision.ACCEPTED:
        raise Phase03ScienceError("scientific handoff requires an accepted deposit-model review")

    georef_bindings = tuple(
        sorted(
            (
                GeoreferenceAcceptanceBinding(
                    acceptance_id=item.acceptance_id,
                    acceptance_file_sha256=sha256_file(path),
                    source_sha256=item.source.sha256,
                    derivative_sha256=item.derivative.sha256,
                )
                for path, item in zip(
                    georeference_acceptance_paths,
                    georeferences,
                    strict=True,
                )
            ),
            key=lambda item: item.acceptance_id,
        )
    )
    return Phase03ScientificHandoff.create(
        phase03_source=source.binding,
        boundary_acceptance=ScientificRecordBinding(
            record_id=boundary.acceptance_id,
            file_sha256=sha256_file(Path(boundary_acceptance_path)),
        ),
        georeference_acceptances=georef_bindings,
        evidence_manifest_bindings=selected_bindings,
        accepted_evidence_ids=tuple(sorted(item.record.evidence_id for item in resolved)),
        accepted_evidence_roles=roles,
        deposit_model_assessment=ScientificRecordBinding(
            record_id=assessment.assessment_id,
            file_sha256=assessment_hash,
        ),
        deposit_model_critique=ScientificRecordBinding(
            record_id=critique.critique_id,
            file_sha256=critique_hash,
        ),
        deposit_model_review=ScientificRecordBinding(
            record_id=review.review_id,
            file_sha256=sha256_file(Path(review_path)),
        ),
        unresolved_gaps=tuple(sorted(set(unresolved_gaps))),
        limitations=tuple(sorted(set(limitations))),
        handed_off_by=review.reviewer,
        handoff_authorization_id=review.reviewer_authorization_id,
        handed_off_at=review.reviewed_at,
        decision="accepted-for-phase04-input",
    )


def write_phase03_science_record(
    record: (
        DepositModelAssessment
        | DepositModelCritique
        | DepositModelReview
        | Phase03ScientificHandoff
    ),
    path: Path,
) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    payload = record.model_dump_json(indent=2) + "\n"
    if target.exists():
        if target.read_text(encoding="utf-8") != payload:
            raise Phase03ScienceError("Phase 03 science record already exists and differs")
        return target
    target.write_text(payload, encoding="utf-8", newline="\n")
    return target


def write_deposit_model_assessment_workbook(
    assessment: DepositModelAssessment,
    path: Path,
) -> Path:
    """Render the sealed assessment into the guide-aligned review workbook."""

    from openpyxl import Workbook
    from openpyxl.styles import Font

    target = Path(path)
    if target.exists():
        raise Phase03ScienceError("deposit-model workbook output already exists")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    assert summary is not None
    summary.title = "Assessment"
    summary_rows = (
        ("assessment_id", assessment.assessment_id),
        ("phase03_run_id", assessment.phase03_run_id),
        ("candidate_model", assessment.candidate_model),
        ("draft_score", assessment.draft_score),
        ("confidence_basis", assessment.confidence_basis),
        ("proposing_job_id", assessment.proposing_job_id or ""),
        ("proposing_response_id", assessment.proposing_response_id or ""),
    )
    summary.append(("field", "value"))
    for row in summary_rows:
        summary.append(row)

    scores = workbook.create_sheet("Criterion Scores")
    score_columns = (
        "criterion_id",
        "criterion",
        "max_points",
        "proposed_points",
        "evidence_ids",
        "missing",
        "rationale",
    )
    scores.append(score_columns)
    for item in assessment.criterion_scores:
        scores.append(
            (
                item.criterion_id,
                item.criterion,
                item.max_points,
                item.proposed_points,
                "; ".join(item.evidence_ids),
                item.missing,
                item.rationale,
            )
        )
    scores.append(("total", "Total", 100, assessment.draft_score, "", "", ""))

    groups = workbook.create_sheet("Evidence Groups")
    group_columns = (
        "group_id",
        "relationship",
        "evidence_ids",
        "source_scale",
        "source_date",
        "correlation_basis",
        "counted_once",
        "confidence_basis",
        "limitations",
    )
    groups.append(group_columns)
    for item in assessment.evidence_groups:
        groups.append(
            (
                item.group_id,
                item.relationship.value,
                "; ".join(item.evidence_ids),
                item.source_scale,
                item.source_date,
                item.correlation_basis,
                item.counted_once,
                item.confidence_basis,
                "; ".join(item.limitations),
            )
        )

    gaps = workbook.create_sheet("Gaps and Validation")
    gaps.append(("category", "value"))
    for category, values in (
        ("missing_evidence", assessment.missing_evidence),
        ("limitation", assessment.limitations),
        ("recommended_validation", assessment.recommended_validation),
    ):
        for value in values:
            gaps.append((category, value))

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            values = (len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(values) + 2, 60)
    workbook.save(target)
    return target


def load_deposit_model_assessment(path: Path) -> DepositModelAssessment:
    return _load(path, DepositModelAssessment, "deposit-model assessment")


def load_deposit_model_critique(path: Path) -> DepositModelCritique:
    return _load(path, DepositModelCritique, "deposit-model critique")


def load_deposit_model_review(path: Path) -> DepositModelReview:
    return _load(path, DepositModelReview, "deposit-model review")


def load_phase03_scientific_handoff(path: Path) -> Phase03ScientificHandoff:
    return _load(path, Phase03ScientificHandoff, "Phase 03 scientific handoff")


def verify_phase03_handoff_inputs(
    handoff: Phase03ScientificHandoff,
    *,
    handoff_path: Path,
    runs_root: Path,
    boundary_acceptance_path: Path,
    georeference_acceptance_paths: tuple[Path, ...],
    assessment_path: Path,
    critique_path: Path,
    review_path: Path,
) -> None:
    """Recheck exact files and the sealed Phase 03 source before Phase 04 uses a handoff."""

    loaded = load_phase03_scientific_handoff(handoff_path)
    if loaded != handoff:
        raise Phase03ScienceError("Phase 03 handoff file differs from the supplied record")
    source = resolve_source_phase(
        runs_root,
        "03",
        handoff.phase03_source.source_run_id,
        require_advance=False,
        require_qaqc_passed=True,
    )
    if source.binding != handoff.phase03_source:
        raise Phase03ScienceError("Phase 03 source run changed after scientific handoff")
    boundary = load_boundary_acceptance(boundary_acceptance_path)
    if (
        boundary.acceptance_id != handoff.boundary_acceptance.record_id
        or sha256_file(Path(boundary_acceptance_path)) != handoff.boundary_acceptance.file_sha256
    ):
        raise Phase03ScienceError("boundary acceptance changed after scientific handoff")
    georeferences = tuple(
        load_georeference_acceptance(path) for path in georeference_acceptance_paths
    )
    current_georef = tuple(
        sorted(
            (
                GeoreferenceAcceptanceBinding(
                    acceptance_id=item.acceptance_id,
                    acceptance_file_sha256=sha256_file(path),
                    source_sha256=item.source.sha256,
                    derivative_sha256=item.derivative.sha256,
                )
                for path, item in zip(
                    georeference_acceptance_paths,
                    georeferences,
                    strict=True,
                )
            ),
            key=lambda item: item.acceptance_id,
        )
    )
    if current_georef != handoff.georeference_acceptances:
        raise Phase03ScienceError("georeference acceptance changed after scientific handoff")
    for path, binding, loader, identity_field in (
        (
            assessment_path,
            handoff.deposit_model_assessment,
            load_deposit_model_assessment,
            "assessment_id",
        ),
        (
            critique_path,
            handoff.deposit_model_critique,
            load_deposit_model_critique,
            "critique_id",
        ),
        (
            review_path,
            handoff.deposit_model_review,
            load_deposit_model_review,
            "review_id",
        ),
    ):
        record = loader(path)
        record_id = getattr(record, identity_field)
        if record_id != binding.record_id or sha256_file(path) != binding.file_sha256:
            raise Phase03ScienceError("deposit-model record changed after scientific handoff")


def _load(path: Path, model: type[_StrictModel], description: str):
    try:
        data = json.loads(Path(path).read_text(encoding="utf-8"), object_pairs_hook=_unique_object)
        return model.model_validate(data)
    except (OSError, UnicodeError, ValueError) as exc:
        raise Phase03ScienceError(f"{description} is invalid") from exc


def _unique_object(pairs: list[tuple[str, object]]) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"duplicate JSON key: {key}")
        result[key] = value
    return result
