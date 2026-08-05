from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

import pytest

from buduunkhad.ai.fingerprint import sha256_file
from buduunkhad.core.boundary_validation import (
    BoundaryAcceptanceRecord,
    BoundFileIdentity,
    write_boundary_authority_record,
)
from buduunkhad.core.evidence_manifest import (
    EvidenceExecutionMode,
    EvidenceLifecycleState,
    EvidenceOrigin,
    EvidenceRecord,
    EvidenceRole,
    EvidenceSourceKind,
    ResolvedEvidence,
)
from buduunkhad.core.execution_policy import ExecutionMode
from buduunkhad.core.georeference import (
    GeoreferenceAcceptanceRecord,
    GeoreferenceFileIdentity,
    write_georeference_record,
)
from buduunkhad.core.phase03_science import (
    DEPOSIT_MODEL_SCORING_CRITERIA,
    CritiqueConclusion,
    CritiqueOrigin,
    DepositModelAssessment,
    DepositModelCriterionScore,
    DepositModelEvidenceGroup,
    EvidenceRelationship,
    Phase03ScienceError,
    ScientificReviewDecision,
    create_deposit_model_assessment,
    create_deposit_model_critique,
    create_deposit_model_review,
    load_deposit_model_assessment,
    load_phase03_scientific_handoff,
    resolve_phase03_scientific_handoff,
    write_deposit_model_assessment_workbook,
    write_phase03_science_record,
)
from buduunkhad.core.run_storage import ResolvedSourcePhase, SourcePhaseBinding


class _Resolver:
    def __init__(self, resolved: tuple[ResolvedEvidence, ...]) -> None:
        self.resolved = resolved

    def resolve_selected(
        self,
        manifest_ids: tuple[str, ...] | list[str],
    ) -> tuple[ResolvedEvidence, ...]:
        assert manifest_ids == ("a" * 64,)
        return self.resolved


def _resolved_evidence(tmp_path: Path) -> tuple[ResolvedEvidence, ...]:
    definitions = (
        (
            "EV-GEOLOGY",
            "geology_units_50k_polygon",
            "geology_units_50k_polygon",
            EvidenceRole.GEOLOGY,
        ),
        (
            "EV-STRUCTURE",
            "faults_structures_line",
            "faults_structures_line",
            EvidenceRole.STRUCTURE,
        ),
        (
            "EV-OCCURRENCE",
            "mineral_occurrences_point",
            "mineral_occurrences_point",
            EvidenceRole.OCCURRENCE,
        ),
        (
            "EV-ALTERATION",
            "ai_accepted_alteration_zones_polygon",
            None,
            EvidenceRole.ALTERATION_SUPPORT,
        ),
    )
    resolved = []
    for evidence_id, layer, target, role in definitions:
        artifact = tmp_path / f"{evidence_id}.gpkg"
        artifact.write_bytes(evidence_id.encode("ascii"))
        record = EvidenceRecord(
            evidence_id=evidence_id,
            source_kind=EvidenceSourceKind.PHASE03_PROMOTION,
            source_run_id="phase03-run",
            source_authority_path=f"reviews/{evidence_id}.promotion-ledger.jsonl",
            source_authority_sha256="b" * 64,
            source_record_id=f"audit-{evidence_id}",
            artifact_path=f"reviews/{evidence_id}.gpkg",
            artifact_sha256=sha256_file(artifact),
            artifact_size_bytes=artifact.stat().st_size,
            layer_name=layer,
            target_layer_name=target,
            evidence_role=role,
            origin=EvidenceOrigin.PHASE03_AI_HANDOFF,
            lifecycle_state=EvidenceLifecycleState.ACCEPTED_EVIDENCE,
            review_record_id=f"audit-{evidence_id}",
            reviewers=("Project Geologist",),
            reviewed_at=datetime(2026, 7, 27, 5, 0, tzinfo=UTC),
            eligible_phases=("03",),
            eligible_modes=(EvidenceExecutionMode.SUPPORT_EVIDENCE,),
            authoritative_for_phase04=False,
            limitations=("Support evidence pending scientific handoff.",),
        )
        resolved.append(
            ResolvedEvidence(
                manifest_id="a" * 64,
                manifest_sha256="c" * 64,
                catalog_entry_id="d" * 64,
                record=record,
                artifact=artifact,
            )
        )
    return tuple(sorted(resolved, key=lambda item: item.record.evidence_id))


def _groups(resolved: tuple[ResolvedEvidence, ...]) -> tuple[DepositModelEvidenceGroup, ...]:
    return tuple(
        DepositModelEvidenceGroup(
            group_id=f"group-{item.record.evidence_role.value}",
            relationship=EvidenceRelationship.SUPPORTING,
            evidence_ids=(item.record.evidence_id,),
            source_scale="source-specific",
            source_date="recorded-in-source",
            correlation_basis=f"One {item.record.evidence_role.value} evidence family.",
            confidence_basis="Human-reviewed accepted evidence.",
        )
        for item in sorted(resolved, key=lambda item: item.record.evidence_role.value)
    )


def _criterion_scores(
    resolved: tuple[ResolvedEvidence, ...],
) -> tuple[DepositModelCriterionScore, ...]:
    evidence_by_role = {item.record.evidence_role: item.record.evidence_id for item in resolved}
    assignments = {
        "favorable-geology": (EvidenceRole.GEOLOGY, 15.0),
        "structure-contact": (EvidenceRole.STRUCTURE, 10.0),
        "known-occurrence": (EvidenceRole.OCCURRENCE, 12.0),
        "remote-sensing-alteration": (EvidenceRole.ALTERATION_SUPPORT, 8.0),
    }
    scores = []
    for criterion_id, criterion, maximum in DEPOSIT_MODEL_SCORING_CRITERIA:
        assignment = assignments.get(criterion_id)
        evidence_id = evidence_by_role.get(assignment[0]) if assignment else None
        points = assignment[1] if assignment and evidence_id else 0.0
        scores.append(
            DepositModelCriterionScore(
                criterion_id=criterion_id,
                criterion=criterion,
                max_points=maximum,
                proposed_points=points,
                evidence_ids=(evidence_id,) if evidence_id else (),
                rationale=(
                    "Exact accepted evidence supports a preliminary partial score."
                    if evidence_id
                    else "No accepted evidence was selected for this criterion."
                ),
                missing=evidence_id is None,
            )
        )
    return tuple(scores)


def _model_chain(tmp_path: Path, resolved: tuple[ResolvedEvidence, ...]):
    assessment = create_deposit_model_assessment(
        phase03_run_id="phase03-run",
        resolver=_Resolver(resolved),
        evidence_manifest_ids=("a" * 64,),
        candidate_model="Porphyry Cu-Au",
        evidence_groups=_groups(resolved),
        missing_evidence=("Field confirmation",),
        confidence_basis="Multiple independent accepted evidence roles.",
        limitations=("Desktop interpretation is not proof of mineralization.",),
        recommended_validation=("Qualified field validation",),
        criterion_scores=_criterion_scores(resolved),
        proposing_job_id="proposal-job",
        proposing_response_id="proposal-response",
    )
    assessment_path = tmp_path / "assessment.json"
    write_phase03_science_record(assessment, assessment_path)
    critique = create_deposit_model_critique(
        assessment_path,
        critic="Independent Critic",
        origin=CritiqueOrigin.AI_INDEPENDENT,
        critiqued_at=datetime(2026, 7, 27, 6, 0, tzinfo=UTC),
        critique_job_id="critique-job",
        critique_response_id="critique-response",
        conclusion=CritiqueConclusion.CHALLENGED,
        findings=("Field confirmation remains missing.",),
    )
    critique_path = tmp_path / "critique.json"
    write_phase03_science_record(critique, critique_path)
    review = create_deposit_model_review(
        assessment_path,
        critique_path,
        reviewer="Project Geologist",
        reviewer_authorization_id="geologist-authorization",
        reviewed_at=datetime(2026, 7, 27, 7, 0, tzinfo=UTC),
        decision=ScientificReviewDecision.ACCEPTED,
        accepted_model="Porphyry Cu-Au",
        accepted_confidence="preliminary-moderate",
        rationale="Accepted as a desktop exploration model with explicit limitations.",
    )
    review_path = tmp_path / "review.json"
    write_phase03_science_record(review, review_path)
    return assessment_path, critique_path, review_path


def test_model_chain_separates_proposal_critique_and_geologist_decision(tmp_path: Path) -> None:
    resolved = _resolved_evidence(tmp_path)
    assessment_path, critique_path, review_path = _model_chain(tmp_path, resolved)

    assert assessment_path.is_file()
    assert critique_path.is_file()
    assert review_path.is_file()
    assessment = load_deposit_model_assessment(assessment_path)
    assert assessment.format_version == "1.1.0"
    assert len(assessment.criterion_scores) == 8
    assert assessment.draft_score == 45
    workbook_path = tmp_path / "assessment.xlsx"
    write_deposit_model_assessment_workbook(assessment, workbook_path)
    import openpyxl

    workbook = openpyxl.load_workbook(workbook_path, read_only=True)
    assert workbook.sheetnames == [
        "Assessment",
        "Criterion Scores",
        "Evidence Groups",
        "Gaps and Validation",
    ]
    score_sheet = workbook["Criterion Scores"]
    assert score_sheet.max_row == 10
    with pytest.raises(Phase03ScienceError, match="separate job identity"):
        create_deposit_model_critique(
            assessment_path,
            critic="Critic",
            origin=CritiqueOrigin.AI_INDEPENDENT,
            critiqued_at=datetime(2026, 7, 27, 6, 0, tzinfo=UTC),
            critique_job_id="proposal-job",
            critique_response_id="another-response",
            conclusion=CritiqueConclusion.SUPPORTED,
            findings=("No new finding.",),
        )


def test_legacy_deposit_model_assessment_remains_readable(tmp_path: Path) -> None:
    resolved = _resolved_evidence(tmp_path)
    assessment_path, _critique_path, _review_path = _model_chain(tmp_path, resolved)
    current = load_deposit_model_assessment(assessment_path)
    values = current.model_dump(
        mode="python",
        exclude={"assessment_id", "criterion_scores"},
    )
    values["format_version"] = "1.0.0"
    legacy = DepositModelAssessment.create(**values)
    legacy_path = tmp_path / "legacy-assessment.json"
    write_phase03_science_record(legacy, legacy_path)

    loaded = load_deposit_model_assessment(legacy_path)
    assert loaded.format_version == "1.0.0"
    assert loaded.criterion_scores == ()


def test_phase03_handoff_resolves_exact_reviews_and_required_roles(
    tmp_path: Path,
    monkeypatch,
) -> None:
    resolved = _resolved_evidence(tmp_path)
    assessment_path, critique_path, review_path = _model_chain(tmp_path, resolved)

    boundary = BoundaryAcceptanceRecord.create(
        validation_id="1" * 64,
        validation_file_sha256="2" * 64,
        processing_run_id="phase01-run",
        source_artifact=BoundFileIdentity(path="source.kmz", sha256="3" * 64, size_bytes=1),
        boundary_derivative=BoundFileIdentity(
            path="boundary.gpkg",
            sha256="4" * 64,
            size_bytes=1,
            layer_name="license_boundary",
        ),
        buffer_derivative=BoundFileIdentity(
            path="buffers.gpkg",
            sha256="5" * 64,
            size_bytes=1,
            layer_name="project_buffers",
        ),
        accepted_attestation_ids=("6" * 64, "7" * 64),
        accepted_at=datetime(2026, 7, 27, 5, 0, tzinfo=UTC),
    )
    boundary_path = tmp_path / "boundary-acceptance.json"
    write_boundary_authority_record(boundary, boundary_path)

    georef = GeoreferenceAcceptanceRecord.create(
        record_id="8" * 64,
        record_file_sha256="9" * 64,
        processing_run_id="phase03-run",
        source=GeoreferenceFileIdentity(path="scan.tif", sha256="a" * 64, size_bytes=1),
        derivative=GeoreferenceFileIdentity(
            path="scan-georef.tif",
            sha256="b" * 64,
            size_bytes=1,
        ),
        accepted_attestation_ids=("c" * 64, "d" * 64),
        accepted_at=datetime(2026, 7, 27, 5, 0, tzinfo=UTC),
    )
    georef_path = tmp_path / "georef-acceptance.json"
    write_georeference_record(georef, georef_path)

    binding = SourcePhaseBinding(
        phase_id="03",
        source_run_id="phase03-run",
        source_manifest_sha256="e" * 64,
        source_phase_sha256="f" * 64,
    )
    source = ResolvedSourcePhase(
        binding=binding,
        phase_dir=tmp_path / "phase03",
        output_artifacts=(),
        sealed_files=(),
        gate_status="blocked",
        gate_provisional=False,
        execution_mode=ExecutionMode.SUPPORT_EVIDENCE,
    )
    monkeypatch.setattr(
        "buduunkhad.core.phase03_science.resolve_source_phase",
        lambda *args, **kwargs: source,
    )
    monkeypatch.setattr(
        "buduunkhad.core.phase03_science.EvidenceAuthorityResolver.resolve_selected",
        lambda self, manifest_ids: resolved,
    )
    runs_root = tmp_path / "runs"
    evidence_root = tmp_path / "evidence"
    runs_root.mkdir()
    evidence_root.mkdir()

    handoff = resolve_phase03_scientific_handoff(
        runs_root=runs_root,
        evidence_root=evidence_root,
        target_epsg=32647,
        phase03_run_id="phase03-run",
        boundary_acceptance_path=boundary_path,
        georeference_acceptance_paths=(georef_path,),
        evidence_manifest_ids=("a" * 64,),
        assessment_path=assessment_path,
        critique_path=critique_path,
        review_path=review_path,
        unresolved_gaps=("Field confirmation",),
        limitations=("Desktop evidence remains support evidence.",),
    )
    output = tmp_path / "phase03-handoff.json"
    write_phase03_science_record(handoff, output)

    assert load_phase03_scientific_handoff(output) == handoff
    assert handoff.readiness_ids_resolved == (
        "METH-READY-004",
        "METH-READY-005",
        "METH-READY-006",
    )
    assert set(handoff.accepted_evidence_roles) >= {
        EvidenceRole.GEOLOGY,
        EvidenceRole.STRUCTURE,
        EvidenceRole.OCCURRENCE,
        EvidenceRole.ALTERATION_SUPPORT,
    }


def test_model_assessment_rejects_double_counted_evidence(tmp_path: Path) -> None:
    resolved = _resolved_evidence(tmp_path)
    duplicate = resolved[0].record.evidence_id
    groups = (
        DepositModelEvidenceGroup(
            group_id="group-a",
            relationship=EvidenceRelationship.SUPPORTING,
            evidence_ids=(duplicate,),
            source_scale="1:50,000",
            source_date="unknown",
            correlation_basis="same derivative family",
            confidence_basis="reviewed",
        ),
        DepositModelEvidenceGroup(
            group_id="group-b",
            relationship=EvidenceRelationship.SUPPORTING,
            evidence_ids=(duplicate,),
            source_scale="1:50,000",
            source_date="unknown",
            correlation_basis="same derivative family",
            confidence_basis="reviewed",
        ),
    )

    with pytest.raises(ValueError, match="selected evidence"):
        create_deposit_model_assessment(
            phase03_run_id="phase03-run",
            resolver=_Resolver(resolved),
            evidence_manifest_ids=("a" * 64,),
            candidate_model="Porphyry Cu-Au",
            evidence_groups=groups,
            missing_evidence=(),
            confidence_basis="reviewed",
            limitations=("support evidence only",),
            recommended_validation=("field validation",),
            criterion_scores=_criterion_scores(resolved),
        )


def test_model_assessment_requires_exact_eight_criterion_rubric(tmp_path: Path) -> None:
    resolved = _resolved_evidence(tmp_path)
    with pytest.raises(ValueError, match="exact eight-criterion rubric"):
        create_deposit_model_assessment(
            phase03_run_id="phase03-run",
            resolver=_Resolver(resolved),
            evidence_manifest_ids=("a" * 64,),
            candidate_model="Porphyry Cu-Au",
            evidence_groups=_groups(resolved),
            missing_evidence=("Field confirmation",),
            confidence_basis="reviewed",
            limitations=("support evidence only",),
            recommended_validation=("field validation",),
            criterion_scores=_criterion_scores(resolved)[:-1],
        )
